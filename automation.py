from openpyxl import Workbook, load_workbook
import sys
wb = load_workbook('data/testsuite/Testsuite.xlsx')
sh=wb['TestSuite'] 
test_folders_loc=[]
for sh_rowB in sh['B']:
        test_folders_loc.append(sh_rowB.value)
print(test_folders_loc)
test_suite_val1=[]
for sh_rowA in sh['A']:
        test_suite_val1.append(sh_rowA.value)
print(test_suite_val1)
input_file_name=list(map(str, sys.argv[1].split(',')))
print("List of files: ", input_file_name)
input_product_name=sys.argv[2]
for file in input_file_name:
   print(file)
   if(file in test_folders_loc):
    print("Exists")
    print(test_folders_loc.index(file))
    t1=test_folders_loc.index(file)+1
    sh['A'+str(t1)].value=1
    print(sh['A'+str(t1)].value)
    wb.save('data/testsuite/Testsuite.xlsx')
    wb2=load_workbook(file)
    sh2=wb2['Testcases']
    product_names=[]
    for sh2_rowC in sh2['C']:
        product_names.append(sh2_rowC.value)
    print(product_names)   
    for i in range(len(product_names)):
        if(product_names[i]==input_product_name):
         print(product_names[i])
         print(i)
         print(type(i))
         sh2['A'+str(i+1)].value=1
         wb2.save(file)
   else:
    print('Input file not exists in list')
